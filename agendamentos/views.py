from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Count, Sum, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
import io
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .forms import AgendamentoForm, TrajetoFormSet, TrajetoFormSetEdit
from .models import Agendamento, Trajeto
from cursos.models import Curso
from veiculos.models import Veiculo


def is_administrador(user):
    """Verifica se o usuário é administrador"""
    return user.is_administrador()


@login_required
def lista_agendamentos(request):
    """Lista agendamentos do usuário (ou todos se admin)"""
    if request.user.is_administrador():
        agendamentos_list = Agendamento.objects.all().select_related(
            'curso', 'professor', 'veiculo')
    else:
        agendamentos_list = Agendamento.objects.filter(
            professor=request.user).select_related('curso', 'veiculo')

    # Filtros
    status = request.GET.get('status')
    if status:
        agendamentos_list = agendamentos_list.filter(status=status)

    curso_id = request.GET.get('curso')
    if curso_id:
        agendamentos_list = agendamentos_list.filter(curso_id=curso_id)

    # Ordenação
    agendamentos_list = agendamentos_list.order_by('-data_inicio')

    # Paginação
    paginator = Paginator(agendamentos_list, 10)  # 10 agendamentos por página
    page = request.GET.get('page')
    
    try:
        agendamentos = paginator.page(page)
    except PageNotAnInteger:
        # Se a página não é um inteiro, mostra a primeira página
        agendamentos = paginator.page(1)
    except EmptyPage:
        # Se a página está fora do range, mostra a última página
        agendamentos = paginator.page(paginator.num_pages)

    # Dados para filtros
    cursos_disponiveis = Curso.objects.filter(ativo=True)

    return render(request, 'agendamentos/lista.html', {
        'agendamentos': agendamentos,
        'status_filter': status,
        'curso_filter': curso_id,
        'cursos_disponiveis': cursos_disponiveis,
    })


@login_required
def criar_agendamento(request):
    """Cria um novo agendamento"""
    if request.method == 'POST':
        form = AgendamentoForm(request.POST, user=request.user)
        formset = TrajetoFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # Valida se pelo menos um trajeto foi preenchido
            trajetos_preenchidos = sum(
                1 for f in formset if f.cleaned_data and
                not f.cleaned_data.get('DELETE', False)
            )
            if trajetos_preenchidos < 1:
                messages.error(
                    request,
                    'Adicione pelo menos um trajeto ao agendamento.'
                )
            else:
                # Calcula total de KM dos trajetos para validação
                total_km_trajetos = sum(
                    f.cleaned_data.get('quilometragem', 0)
                    for f in formset if f.cleaned_data and
                    not f.cleaned_data.get('DELETE', False)
                )
                
                # Valida o limite de KM manualmente
                try:
                    form.validar_limite_km_manual(total_km_trajetos)
                    
                    with transaction.atomic():
                        agendamento = form.save(commit=False)
                        agendamento.professor = request.user
                        agendamento.status = 'pendente'
                        agendamento.save()

                        # Salva os trajetos
                        formset.instance = agendamento
                        formset.save()

                        messages.success(
                            request,
                            'Agendamento criado com sucesso! '
                            'Aguarde a aprovação do administrador.'
                        )
                        return redirect('agendamentos:lista')
                except ValidationError as e:
                    # Extrai a mensagem de erro corretamente
                    if hasattr(e, 'message'):
                        error_msg = e.message
                    elif hasattr(e, 'messages') and e.messages:
                        error_msg = ' '.join(e.messages)
                    else:
                        error_msg = str(e)
                    messages.error(request, error_msg)
    else:
        form = AgendamentoForm(user=request.user)
        formset = TrajetoFormSet()

    return render(request, 'agendamentos/form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Novo Agendamento'
    })


@login_required
def editar_agendamento(request, pk):
    """Edita um agendamento existente"""
    agendamento = get_object_or_404(Agendamento, pk=pk)

    # Verifica permissão: apenas o professor dono ou admin pode editar
    if not request.user.is_administrador() and agendamento.professor != request.user:
        messages.error(
            request, 'Você não tem permissão para editar este agendamento.')
        return redirect('agendamentos:lista')

    # Não permite editar agendamentos aprovados ou reprovados (apenas pendentes)
    if agendamento.status != 'pendente':
        messages.warning(
            request, 'Apenas agendamentos pendentes podem ser editados.')
        return redirect('agendamentos:lista')

    if request.method == 'POST':
        form = AgendamentoForm(
            request.POST, instance=agendamento, user=request.user)
        formset = TrajetoFormSetEdit(request.POST, instance=agendamento)

        if form.is_valid() and formset.is_valid():
            # Calcula total de KM dos trajetos para validação
            total_km_trajetos = sum(
                f.cleaned_data.get('quilometragem', 0)
                for f in formset if f.cleaned_data and
                not f.cleaned_data.get('DELETE', False)
            )
            
            # Valida o limite de KM manualmente
            try:
                form.validar_limite_km_manual(total_km_trajetos)
                
                with transaction.atomic():
                    agendamento = form.save()
                    formset.save()

                    messages.success(
                        request, 'Agendamento atualizado com sucesso!')
                    return redirect('agendamentos:lista')
            except ValidationError as e:
                # Extrai a mensagem de erro corretamente
                if hasattr(e, 'message'):
                    error_msg = e.message
                elif hasattr(e, 'messages') and e.messages:
                    error_msg = ' '.join(e.messages)
                else:
                    error_msg = str(e)
                messages.error(request, error_msg)
    else:
        form = AgendamentoForm(instance=agendamento, user=request.user)
        formset = TrajetoFormSetEdit(instance=agendamento)

    return render(request, 'agendamentos/form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Editar Agendamento',
        'agendamento': agendamento
    })


@login_required
def detalhe_agendamento(request, pk):
    """Exibe detalhes de um agendamento"""
    agendamento = get_object_or_404(Agendamento, pk=pk)

    # Verifica permissão
    if not request.user.is_administrador() and agendamento.professor != request.user:
        messages.error(
            request, 'Você não tem permissão para ver este agendamento.')
        return redirect('agendamentos:lista')

    trajetos = agendamento.trajetos.all()
    total_km = agendamento.get_total_km()

    return render(request, 'agendamentos/detalhe.html', {
        'agendamento': agendamento,
        'trajetos': trajetos,
        'total_km': total_km
    })


@login_required
def deletar_agendamento(request, pk):
    """Deleta um agendamento"""
    agendamento = get_object_or_404(Agendamento, pk=pk)

    # Verifica permissão
    if not request.user.is_administrador() and agendamento.professor != request.user:
        messages.error(
            request, 'Você não tem permissão para deletar este agendamento.')
        return redirect('agendamentos:lista')

    # Não permite deletar agendamentos aprovados
    if agendamento.status == 'aprovado':
        messages.warning(
            request, 'Agendamentos aprovados não podem ser deletados.')
        return redirect('agendamentos:lista')

    if request.method == 'POST':
        agendamento.delete()
        messages.success(request, 'Agendamento deletado com sucesso!')
        return redirect('agendamentos:lista')

    return render(request, 'agendamentos/deletar.html', {'agendamento': agendamento})


@login_required
@user_passes_test(is_administrador)
def aprovacao_agendamentos(request):
    """Lista agendamentos pendentes para aprovação"""
    agendamentos_list = Agendamento.objects.filter(
        status='pendente'
    ).select_related('curso', 'professor', 'veiculo').order_by('data_inicio')

    # Paginação
    paginator = Paginator(agendamentos_list, 12)  # 12 agendamentos por página
    page = request.GET.get('page')
    
    try:
        agendamentos_pendentes = paginator.page(page)
    except PageNotAnInteger:
        agendamentos_pendentes = paginator.page(1)
    except EmptyPage:
        agendamentos_pendentes = paginator.page(paginator.num_pages)

    return render(request, 'agendamentos/aprovacao.html', {
        'agendamentos': agendamentos_pendentes
    })


@login_required
@user_passes_test(is_administrador)
def aprovar_agendamento(request, pk):
    """Aprova um agendamento"""
    agendamento = get_object_or_404(Agendamento, pk=pk)

    if request.method == 'POST':
        try:
            agendamento.aprovar()
            messages.success(request, f'Agendamento aprovado com sucesso!')
            return redirect('agendamentos:detalhe', pk=pk)
        except ValidationError as e:
            # Extrai a mensagem de erro corretamente
            if hasattr(e, 'message'):
                error_msg = e.message
            elif hasattr(e, 'messages') and e.messages:
                error_msg = ' '.join(e.messages)
            else:
                error_msg = str(e)
            messages.error(request, f'Erro ao aprovar agendamento: {error_msg}')
            return redirect('agendamentos:detalhe', pk=pk)

    return redirect('agendamentos:detalhe', pk=pk)


@login_required
@user_passes_test(is_administrador)
def reprovar_agendamento(request, pk):
    """Reprova um agendamento"""
    agendamento = get_object_or_404(Agendamento, pk=pk)

    if request.method == 'POST':
        motivo = request.POST.get('motivo', '')
        if not motivo:
            messages.error(
                request, 'É necessário informar o motivo da reprovação.')
            return redirect('agendamentos:reprovar', pk=pk)

        agendamento.reprovar(motivo)
        messages.success(request, 'Agendamento reprovado.')
        return redirect('agendamentos:detalhe', pk=pk)

    return render(request, 'agendamentos/reprovar.html', {'agendamento': agendamento})


@login_required
def agendamentos_json(request):
    """Retorna agendamentos em formato JSON para o calendário"""
    if request.user.is_administrador():
        agendamentos = Agendamento.objects.all()
    else:
        agendamentos = Agendamento.objects.filter(professor=request.user)

    eventos = []
    for agendamento in agendamentos:
        # Define cor baseada no status
        if agendamento.status == 'pendente':
            color = '#ffc107'  # Amarelo
        elif agendamento.status == 'aprovado':
            color = '#28a745'  # Verde
        else:  # reprovado
            color = '#dc3545'  # Vermelho

        eventos.append({
            'id': agendamento.id,
            'title': f"{agendamento.curso.nome} - {agendamento.veiculo.placa}",
            'start': agendamento.data_inicio.isoformat(),
            'end': agendamento.data_fim.isoformat(),
            'color': color,
            'url': f'/agendamentos/{agendamento.id}/',
        })

    return JsonResponse(eventos, safe=False)


@login_required
@user_passes_test(is_administrador)
def relatorio_geral(request):
    """Relatório geral do sistema de agendamentos"""
    # Obter filtros da URL
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')
    curso_id = request.GET.get('curso')
    status = request.GET.get('status')
    
    # Data atual para valores padrão
    hoje = timezone.now()
    
    # Se não especificado, usar ano e mês atuais
    if not ano:
        ano = hoje.year
    else:
        ano = int(ano)
        
    if not mes:
        mes = hoje.month
    else:
        mes = int(mes)
    
    # Query base
    agendamentos = Agendamento.objects.filter(
        data_inicio__year=ano,
        data_inicio__month=mes
    ).select_related('curso', 'professor', 'veiculo')
    
    # Aplicar filtros
    if curso_id:
        agendamentos = agendamentos.filter(curso_id=curso_id)
    
    if status:
        agendamentos = agendamentos.filter(status=status)
    
    # Estatísticas gerais
    total_agendamentos = agendamentos.count()
    agendamentos_aprovados = agendamentos.filter(status='aprovado').count()
    agendamentos_pendentes = agendamentos.filter(status='pendente').count()
    agendamentos_reprovados = agendamentos.filter(status='reprovado').count()
    
    # Estatísticas por status
    stats_status = {
        'aprovado': agendamentos_aprovados,
        'pendente': agendamentos_pendentes,
        'reprovado': agendamentos_reprovados,
        'total': total_agendamentos
    }
    
    # KM total por curso
    cursos_km = {}
    total_km = 0
    
    for agendamento in agendamentos.filter(status='aprovado'):
        curso_nome = agendamento.curso.nome
        km_agendamento = agendamento.get_total_km()
        
        if curso_nome not in cursos_km:
            cursos_km[curso_nome] = {
                'km_total': 0,
                'agendamentos': 0,
                'limite_mensal': agendamento.curso.limite_km_mensal
            }
        
        cursos_km[curso_nome]['km_total'] += km_agendamento
        cursos_km[curso_nome]['agendamentos'] += 1
        total_km += km_agendamento
    
    # Calcular percentual de uso do limite para cada curso
    for curso_nome, dados in cursos_km.items():
        dados['percentual_uso'] = (dados['km_total'] / dados['limite_mensal']) * 100
        dados['km_disponivel'] = dados['limite_mensal'] - dados['km_total']
    
    # Agendamentos por veículo
    veiculos_stats = agendamentos.values(
        'veiculo__placa', 'veiculo__marca', 'veiculo__modelo'
    ).annotate(
        total_agendamentos=Count('id')
    ).order_by('-total_agendamentos')
    
    # Agendamentos por professor
    professores_stats = agendamentos.values(
        'professor__first_name', 'professor__last_name'
    ).annotate(
        total_agendamentos=Count('id')
    ).order_by('-total_agendamentos')
    
    # Dados para os filtros
    anos_disponiveis = list(range(2023, hoje.year + 2))
    meses_disponiveis = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
    ]
    cursos_disponiveis = Curso.objects.filter(ativo=True)
    
    # Nome do mês atual
    nome_mes = dict(meses_disponiveis)[mes]
    
    # Paginação dos agendamentos
    agendamentos_list = agendamentos.order_by('-data_inicio')
    paginator = Paginator(agendamentos_list, 15)  # 15 agendamentos por página
    page = request.GET.get('page')
    
    try:
        agendamentos_paginados = paginator.page(page)
    except PageNotAnInteger:
        agendamentos_paginados = paginator.page(1)
    except EmptyPage:
        agendamentos_paginados = paginator.page(paginator.num_pages)
    
    context = {
        'stats_status': stats_status,
        'cursos_km': cursos_km,
        'total_km': total_km,
        'veiculos_stats': veiculos_stats,
        'professores_stats': professores_stats,
        'agendamentos': agendamentos_paginados,
        
        # Filtros atuais
        'ano_atual': ano,
        'mes_atual': mes,
        'nome_mes': nome_mes,
        'curso_atual': curso_id,
        'status_atual': status,
        
        # Opções para filtros
        'anos_disponiveis': anos_disponiveis,
        'meses_disponiveis': meses_disponiveis,
        'cursos_disponiveis': cursos_disponiveis,
        'status_choices': Agendamento.STATUS_CHOICES,
    }
    
    return render(request, 'agendamentos/relatorio_geral.html', context)


@login_required
@user_passes_test(is_administrador)
def relatorio_por_curso(request):
    """Relatório detalhado por curso"""
    curso_id = request.GET.get('curso')
    ano = request.GET.get('ano')
    
    hoje = timezone.now()
    if not ano:
        ano = hoje.year
    else:
        ano = int(ano)
    
    # Se não especificar curso, pegar o primeiro ativo
    if not curso_id:
        primeiro_curso = Curso.objects.filter(ativo=True).first()
        if primeiro_curso:
            curso_id = primeiro_curso.id
    
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Dados por mês do ano
    dados_mensais = []
    total_km_ano = 0
    
    for mes_num in range(1, 13):
        agendamentos_mes = Agendamento.objects.filter(
            curso=curso,
            status='aprovado',
            data_inicio__year=ano,
            data_inicio__month=mes_num
        )
        
        km_mes = sum(a.get_total_km() for a in agendamentos_mes)
        total_agendamentos_mes = agendamentos_mes.count()
        
        dados_mensais.append({
            'mes_numero': mes_num,
            'mes_nome': calendar.month_name[mes_num],
            'km_utilizados': km_mes,
            'agendamentos': total_agendamentos_mes,
            'percentual_limite': (km_mes / curso.limite_km_mensal) * 100 if curso.limite_km_mensal > 0 else 0,
            'km_disponiveis': curso.limite_km_mensal - km_mes
        })
        
        total_km_ano += km_mes
    
    # Agendamentos detalhados do ano
    agendamentos_ano = Agendamento.objects.filter(
        curso=curso,
        data_inicio__year=ano
    ).select_related('professor', 'veiculo').order_by('-data_inicio')
    
    # Estatísticas do curso no ano
    stats_ano = {
        'total_km': total_km_ano,
        'limite_anual': curso.limite_km_mensal * 12,
        'percentual_uso_anual': (total_km_ano / (curso.limite_km_mensal * 12)) * 100 if curso.limite_km_mensal > 0 else 0,
        'total_agendamentos': agendamentos_ano.count(),
        'agendamentos_aprovados': agendamentos_ano.filter(status='aprovado').count(),
        'agendamentos_pendentes': agendamentos_ano.filter(status='pendente').count(),
        'agendamentos_reprovados': agendamentos_ano.filter(status='reprovado').count(),
    }
    
    context = {
        'curso': curso,
        'ano': ano,
        'dados_mensais': dados_mensais,
        'agendamentos_ano': agendamentos_ano,
        'stats_ano': stats_ano,
        'cursos_disponiveis': Curso.objects.filter(ativo=True),
        'anos_disponiveis': list(range(2023, hoje.year + 2)),
    }
    
    return render(request, 'agendamentos/relatorio_por_curso.html', context)


@login_required
@user_passes_test(is_administrador)
def exportar_relatorio_excel(request):
    """Exporta relatório geral em Excel"""
    # Obter filtros
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')
    curso_id = request.GET.get('curso')
    status = request.GET.get('status')
    
    hoje = timezone.now()
    if not ano:
        ano = hoje.year
    else:
        ano = int(ano)
        
    if not mes:
        mes = hoje.month
    else:
        mes = int(mes)
    
    # Query base
    agendamentos = Agendamento.objects.filter(
        data_inicio__year=ano,
        data_inicio__month=mes
    ).select_related('curso', 'professor', 'veiculo')
    
    # Aplicar filtros
    if curso_id:
        agendamentos = agendamentos.filter(curso_id=curso_id)
    if status:
        agendamentos = agendamentos.filter(status=status)
    
    # Criar arquivo Excel
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Formatos
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'align': 'center',
        'border': 1
    })
    
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'left'
    })
    
    number_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '#,##0'
    })
    
    # Aba Principal - Agendamentos
    worksheet = workbook.add_worksheet('Agendamentos')
    
    # Título
    meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    titulo = f'Relatório de Agendamentos - {meses[mes]} {ano}'
    worksheet.merge_range('A1:H1', titulo, title_format)
    
    # Cabeçalhos
    headers = ['Data Início', 'Data Fim', 'Curso', 'Professor', 'Veículo', 'Status', 'KM Total', 'Observações']
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, header_format)
    
    # Dados
    row = 3
    for agendamento in agendamentos.order_by('-data_inicio'):
        worksheet.write(row, 0, agendamento.data_inicio.strftime('%d/%m/%Y %H:%M'), cell_format)
        worksheet.write(row, 1, agendamento.data_fim.strftime('%d/%m/%Y %H:%M'), cell_format)
        worksheet.write(row, 2, agendamento.curso.nome, cell_format)
        worksheet.write(row, 3, agendamento.professor.get_full_name(), cell_format)
        worksheet.write(row, 4, f"{agendamento.veiculo.placa} - {agendamento.veiculo.marca} {agendamento.veiculo.modelo}", cell_format)
        worksheet.write(row, 5, agendamento.get_status_display(), cell_format)
        worksheet.write(row, 6, agendamento.get_total_km(), number_format)
        worksheet.write(row, 7, agendamento.observacoes or '', cell_format)
        row += 1
    
    # Ajustar largura das colunas
    worksheet.set_column('A:A', 15)  # Data Início
    worksheet.set_column('B:B', 15)  # Data Fim
    worksheet.set_column('C:C', 20)  # Curso
    worksheet.set_column('D:D', 25)  # Professor
    worksheet.set_column('E:E', 25)  # Veículo
    worksheet.set_column('F:F', 12)  # Status
    worksheet.set_column('G:G', 10)  # KM
    worksheet.set_column('H:H', 30)  # Observações
    
    # Aba Estatísticas por Curso
    stats_worksheet = workbook.add_worksheet('Estatísticas por Curso')
    stats_worksheet.merge_range('A1:F1', f'Estatísticas por Curso - {meses[mes]} {ano}', title_format)
    
    # Cabeçalhos estatísticas
    stats_headers = ['Curso', 'KM Utilizados', 'Limite Mensal', 'KM Disponíveis', '% Uso', 'Agendamentos']
    for col, header in enumerate(stats_headers):
        stats_worksheet.write(2, col, header, header_format)
    
    # Dados por curso
    cursos_km = {}
    for agendamento in agendamentos.filter(status='aprovado'):
        curso_nome = agendamento.curso.nome
        km_agendamento = agendamento.get_total_km()
        
        if curso_nome not in cursos_km:
            cursos_km[curso_nome] = {
                'km_total': 0,
                'agendamentos': 0,
                'limite_mensal': agendamento.curso.limite_km_mensal
            }
        
        cursos_km[curso_nome]['km_total'] += km_agendamento
        cursos_km[curso_nome]['agendamentos'] += 1
    
    row = 3
    for curso_nome, dados in cursos_km.items():
        percentual = (dados['km_total'] / dados['limite_mensal']) * 100 if dados['limite_mensal'] > 0 else 0
        km_disponiveis = dados['limite_mensal'] - dados['km_total']
        
        stats_worksheet.write(row, 0, curso_nome, cell_format)
        stats_worksheet.write(row, 1, dados['km_total'], number_format)
        stats_worksheet.write(row, 2, dados['limite_mensal'], number_format)
        stats_worksheet.write(row, 3, km_disponiveis, number_format)
        stats_worksheet.write(row, 4, f"{percentual:.1f}%", cell_format)
        stats_worksheet.write(row, 5, dados['agendamentos'], number_format)
        row += 1
    
    # Ajustar largura das colunas da aba estatísticas
    stats_worksheet.set_column('A:A', 25)  # Curso
    stats_worksheet.set_column('B:E', 15)  # Números
    stats_worksheet.set_column('F:F', 15)  # Agendamentos
    
    workbook.close()
    output.seek(0)
    
    # Resposta HTTP
    filename = f'relatorio_agendamentos_{meses[mes].lower()}_{ano}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@user_passes_test(is_administrador)
def exportar_relatorio_pdf(request):
    """Exporta relatório geral em PDF"""
    # Obter filtros
    ano = request.GET.get('ano')
    mes = request.GET.get('mes')
    curso_id = request.GET.get('curso')
    status = request.GET.get('status')
    
    hoje = timezone.now()
    if not ano:
        ano = hoje.year
    else:
        ano = int(ano)
        
    if not mes:
        mes = hoje.month
    else:
        mes = int(mes)
    
    # Query base
    agendamentos = Agendamento.objects.filter(
        data_inicio__year=ano,
        data_inicio__month=mes
    ).select_related('curso', 'professor', 'veiculo')
    
    # Aplicar filtros
    if curso_id:
        agendamentos = agendamentos.filter(curso_id=curso_id)
        curso_nome = Curso.objects.get(id=curso_id).nome
    else:
        curso_nome = "Todos os Cursos"
        
    if status:
        agendamentos = agendamentos.filter(status=status)
    
    # Configurar PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Título
    meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    titulo = f'Relatório de Agendamentos<br/>{meses[mes]} {ano}'
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 12))
    
    # Filtros aplicados
    filtros_text = f"<b>Filtros:</b> {curso_nome}"
    if status:
        filtros_text += f" | Status: {dict(Agendamento.STATUS_CHOICES)[status]}"
    elements.append(Paragraph(filtros_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Estatísticas gerais
    total_agendamentos = agendamentos.count()
    aprovados = agendamentos.filter(status='aprovado').count()
    pendentes = agendamentos.filter(status='pendente').count()
    reprovados = agendamentos.filter(status='reprovado').count()
    total_km = sum(a.get_total_km() for a in agendamentos.filter(status='aprovado'))
    
    stats_data = [
        ['Estatísticas Gerais', '', '', ''],
        ['Total de Agendamentos', str(total_agendamentos), 'Aprovados', str(aprovados)],
        ['Pendentes', str(pendentes), 'Reprovados', str(reprovados)],
        ['Total KM (Aprovados)', f'{total_km} km', '', '']
    ]
    
    stats_table = Table(stats_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Tabela de agendamentos
    if agendamentos.exists():
        elements.append(Paragraph('<b>Lista de Agendamentos</b>', styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        # Cabeçalhos
        data = [['Data/Hora', 'Curso', 'Professor', 'Veículo', 'Status', 'KM']]
        
        # Dados
        for agendamento in agendamentos.order_by('-data_inicio')[:50]:  # Limitar a 50 para não sobrecarregar
            data.append([
                agendamento.data_inicio.strftime('%d/%m %H:%M'),
                agendamento.curso.nome[:15] + '...' if len(agendamento.curso.nome) > 15 else agendamento.curso.nome,
                agendamento.professor.get_full_name()[:20] + '...' if len(agendamento.professor.get_full_name()) > 20 else agendamento.professor.get_full_name(),
                agendamento.veiculo.placa,
                agendamento.get_status_display(),
                f'{agendamento.get_total_km()} km'
            ])
        
        table = Table(data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.6*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        if agendamentos.count() > 50:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f'<i>* Mostrando os primeiros 50 de {agendamentos.count()} agendamentos</i>', styles['Normal']))
    
    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Resposta HTTP
    filename = f'relatorio_agendamentos_{meses[mes].lower()}_{ano}.pdf'
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@user_passes_test(is_administrador)
def exportar_curso_excel(request):
    """Exporta relatório por curso em Excel"""
    curso_id = request.GET.get('curso')
    ano = request.GET.get('ano')
    
    hoje = timezone.now()
    if not ano:
        ano = hoje.year
    else:
        ano = int(ano)
    
    if not curso_id:
        primeiro_curso = Curso.objects.filter(ativo=True).first()
        if primeiro_curso:
            curso_id = primeiro_curso.id
    
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Criar arquivo Excel
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Formatos
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'align': 'center'
    })
    
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'align': 'center',
        'border': 1
    })
    
    cell_format = workbook.add_format({'border': 1})
    number_format = workbook.add_format({'border': 1, 'num_format': '#,##0'})
    
    # Aba de dados mensais
    worksheet = workbook.add_worksheet('Dados Mensais')
    worksheet.merge_range('A1:G1', f'{curso.nome} - {ano}', title_format)
    
    # Cabeçalhos
    headers = ['Mês', 'KM Utilizados', 'Limite Mensal', 'KM Disponíveis', '% Uso', 'Agendamentos', 'Status']
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, header_format)
    
    # Dados mensais
    row = 3
    for mes_num in range(1, 13):
        agendamentos_mes = Agendamento.objects.filter(
            curso=curso,
            status='aprovado',
            data_inicio__year=ano,
            data_inicio__month=mes_num
        )
        
        km_mes = sum(a.get_total_km() for a in agendamentos_mes)
        total_agendamentos_mes = agendamentos_mes.count()
        percentual = (km_mes / curso.limite_km_mensal) * 100 if curso.limite_km_mensal > 0 else 0
        km_disponiveis = curso.limite_km_mensal - km_mes
        
        if percentual > 100:
            status = "Excedido"
        elif percentual > 80:
            status = "Atenção"
        else:
            status = "Normal"
        
        worksheet.write(row, 0, calendar.month_name[mes_num], cell_format)
        worksheet.write(row, 1, km_mes, number_format)
        worksheet.write(row, 2, curso.limite_km_mensal, number_format)
        worksheet.write(row, 3, km_disponiveis, number_format)
        worksheet.write(row, 4, f"{percentual:.1f}%", cell_format)
        worksheet.write(row, 5, total_agendamentos_mes, number_format)
        worksheet.write(row, 6, status, cell_format)
        row += 1
    
    # Ajustar larguras
    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:F', 12)
    worksheet.set_column('G:G', 10)
    
    workbook.close()
    output.seek(0)
    
    # Resposta HTTP
    filename = f'relatorio_{curso.nome.lower().replace(" ", "_")}_{ano}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
